"""
SSH 批量执行工具 - 主机管理模块
"""

import csv

import openpyxl
from openpyxl.styles import Font, PatternFill

from config import AppConfig, HostConfig


class HostManager:
    """主机管理器"""

    def __init__(self, config: AppConfig):
        self.config = config

    def add_host(self, host: HostConfig) -> bool:
        """添加主机"""
        if not host.host:
            return False

        key = f"{host.host}:{host.port}"
        if any(f"{h.host}:{h.port}" == key for h in self.config.hosts):
            return False

        # 如果名称为空，使用主机地址
        if not host.name:
            host.name = host.host

        self.config.hosts.append(host)
        return True

    def update_host(self, old_key: str, new_host: HostConfig) -> bool:
        """更新主机"""
        for i, host in enumerate(self.config.hosts):
            if f"{host.host}:{host.port}" == old_key:
                self.config.hosts[i] = new_host
                return True
        return False

    def remove_host(self, key: str) -> bool:
        """删除主机"""
        for i, host in enumerate(self.config.hosts):
            if f"{host.host}:{host.port}" == key:
                del self.config.hosts[i]
                return True
        return False

    def get_host_by_key(self, key: str) -> HostConfig | None:
        """根据key获取主机"""
        for host in self.config.hosts:
            if f"{host.host}:{host.port}" == key:
                return host
        return None

    def get_hosts_by_group(self, group: str) -> list[HostConfig]:
        """按分组获取主机"""
        if group == "全部":
            return self.config.hosts
        return [h for h in self.config.hosts if h.group == group]

    def add_group(self, name: str) -> bool:
        """添加分组"""
        if name and name not in self.config.groups:
            self.config.groups.append(name)
            return True
        return False

    def rename_group(self, old_name: str, new_name: str) -> bool:
        """重命名分组"""
        if old_name not in self.config.groups or new_name in self.config.groups:
            return False

        idx = self.config.groups.index(old_name)
        self.config.groups[idx] = new_name

        # 更新所有该分组下的主机
        for host in self.config.hosts:
            if host.group == old_name:
                host.group = new_name

        return True

    def delete_group(self, name: str) -> bool:
        """删除分组"""
        if name == "默认分组" or name not in self.config.groups:
            return False

        self.config.groups.remove(name)

        # 将该分组下的主机移到默认分组
        for host in self.config.hosts:
            if host.group == name:
                host.group = "默认分组"

        return True

    def import_from_excel(self, file_path: str) -> int:
        """从Excel导入主机列表"""
        count = 0
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers[cell_value.strip().lower()] = col

        # 支持多种列名
        host_aliases = ["host", "主机", "服务器", "ip", "地址"]
        name_aliases = ["name", "名称", "别名", "标识"]
        port_aliases = ["port", "端口", "端口号"]
        username_aliases = ["username", "用户名", "用户", "登录名"]
        password_aliases = ["password", "密码", "口令"]
        switch_to_root_aliases = [
            "switch_to_root",
            "切换root",
            "切换到root",
            "是否切换root",
            "su root",
        ]
        root_password_aliases = [
            "root_password",
            "root密码",
            "root口令",
            "超级用户密码",
        ]

        def get_col(aliases):
            for alias in aliases:
                if alias in headers:
                    return headers[alias]
            return 0

        host_col = get_col(host_aliases)
        if host_col == 0:
            wb.close()
            raise ValueError("Excel文件缺少必需的列：主机/host")

        name_col = get_col(name_aliases)
        port_col = get_col(port_aliases)
        username_col = get_col(username_aliases)
        password_col = get_col(password_aliases)
        switch_to_root_col = get_col(switch_to_root_aliases)
        root_password_col = get_col(root_password_aliases)

        for row in range(2, ws.max_row + 1):
            host_val = ws.cell(row=row, column=host_col).value
            if not host_val:
                continue

            name_val = ws.cell(row=row, column=name_col).value if name_col else ""
            port_val = ws.cell(row=row, column=port_col).value if port_col else 22
            username_val = (
                ws.cell(row=row, column=username_col).value if username_col else ""
            )
            password_val = (
                ws.cell(row=row, column=password_col).value if password_col else ""
            )
            switch_to_root_val = (
                ws.cell(row=row, column=switch_to_root_col).value
                if switch_to_root_col
                else False
            )
            root_password_val = (
                ws.cell(row=row, column=root_password_col).value
                if root_password_col
                else ""
            )

            host = HostConfig(
                name=str(name_val) if name_val else str(host_val),
                host=str(host_val),
                port=int(port_val) if port_val else 22,
                username=str(username_val) if username_val else "",
                password=str(password_val) if password_val else "",
                switch_to_root=str(switch_to_root_val).lower()
                in ["true", "是", "1", "yes"]
                if switch_to_root_val
                else False,
                root_password=str(root_password_val) if root_password_val else "",
            )

            if self.add_host(host):
                count += 1

        wb.close()
        return count

    def import_from_csv(self, file_path: str) -> int:
        """从CSV导入主机列表"""
        count = 0

        with open(file_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)

            # 支持多种列名
            field_mapping = {
                "host": ["host", "主机", "服务器", "ip", "地址"],
                "name": ["name", "名称", "别名", "标识"],
                "port": ["port", "端口", "端口号"],
                "username": ["username", "用户名", "用户", "登录名"],
                "password": ["password", "密码", "口令"],
                "switch_to_root": [
                    "switch_to_root",
                    "切换root",
                    "切换到root",
                    "是否切换root",
                    "su root",
                ],
                "root_password": [
                    "root_password",
                    "root密码",
                    "root口令",
                    "超级用户密码",
                ],
            }

            # 找到正确的列名
            header_map = {}
            for field, aliases in field_mapping.items():
                for alias in aliases:
                    if alias.lower() in [h.lower() for h in reader.fieldnames]:
                        for h in reader.fieldnames:
                            if h.lower() == alias.lower():
                                header_map[field] = h
                                break
                        break

            if "host" not in header_map:
                raise ValueError("CSV文件缺少必需的列：主机/host")

            for row in reader:
                host_val = row.get(header_map["host"], "").strip()
                if not host_val:
                    continue

                name_val = row.get(header_map.get("name"), "").strip()
                port_val = row.get(header_map.get("port"), "22").strip()
                username_val = row.get(header_map.get("username"), "").strip()
                password_val = row.get(header_map.get("password"), "").strip()
                switch_to_root_val = row.get(
                    header_map.get("switch_to_root"), ""
                ).strip()
                root_password_val = row.get(header_map.get("root_password"), "").strip()

                host = HostConfig(
                    name=name_val if name_val else host_val,
                    host=host_val,
                    port=int(port_val) if port_val.isdigit() else 22,
                    username=username_val,
                    password=password_val,
                    switch_to_root=switch_to_root_val.lower()
                    in ["true", "是", "1", "yes"],
                    root_password=root_password_val,
                )

                if self.add_host(host):
                    count += 1

        return count

    def export_to_excel(self, file_path: str) -> bool:
        """导出到Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "主机列表"

            headers = [
                "名称",
                "主机",
                "端口",
                "用户名",
                "密码",
                "认证方式",
                "切换root【填是或否】",
                "root密码",
                "分组",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

            for row, host in enumerate(self.config.hosts, 2):
                ws.cell(row=row, column=1, value=host.name)
                ws.cell(row=row, column=2, value=host.host)
                ws.cell(row=row, column=3, value=host.port)
                ws.cell(row=row, column=4, value=host.username)
                ws.cell(row=row, column=5, value=host.password)
                ws.cell(row=row, column=6, value="密钥" if host.use_key else "密码")
                ws.cell(row=row, column=7, value="是" if host.switch_to_root else "")
                ws.cell(row=row, column=8, value=host.root_password)
                ws.cell(row=row, column=9, value=host.group)

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            wb.save(file_path)
            wb.close()
            return True
        except Exception as e:
            raise ValueError(f"导出失败: {str(e)}")

    def export_to_json(self, file_path: str) -> bool:
        """导出到JSON"""
        import json

        try:
            data = []
            for host in self.config.hosts:
                data.append(
                    {
                        "name": host.name,
                        "host": host.host,
                        "port": host.port,
                        "username": host.username,
                        "password": host.password,
                        "use_key": host.use_key,
                        "key_file": host.key_file,
                        "key_passphrase": host.key_passphrase,
                        "sudo_enabled": host.sudo_enabled,
                        "sudo_password": host.sudo_password,
                        "group": host.group,
                    }
                )

            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return True
        except Exception as e:
            raise ValueError(f"导出失败: {str(e)}")
