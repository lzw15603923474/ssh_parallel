"""
SSH 并行管理工具 - 专业版
功能：
1. 多命令编排执行（顺序/并行、状态显示、依赖关系、日志记录）
2. 用户权限与超时管理（sudo 切换、会话保持、单独超时设置）
3. 主机配置文件导入导出（Excel/JSON、批量选择、自定义字段）
4. 命令安全提醒（单次弹窗、不再提醒、确认执行）
"""

import json
import os
import threading
import time
import tkinter as tk
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl
import paramiko
from openpyxl.styles import Alignment, Font, PatternFill


class CommandExecutionMode(Enum):
    """命令执行模式"""

    SEQUENTIAL = "sequential"
    PARALLEL = "parallel"


class CommandStatus(Enum):
    """命令执行状态"""

    PENDING = "等待中"
    RUNNING = "执行中"
    SUCCESS = "成功"
    FAILED = "失败"
    TIMEOUT = "超时"
    CANCELLED = "已取消"


@dataclass
class ServerConfig:
    """服务器配置"""

    name: str
    host: str
    port: int = 22
    username: str = ""
    password: str = ""
    use_key: bool = False
    key_file: str = ""
    timeout: int = 10
    use_sudo: bool = False
    sudo_user: str = "root"
    sudo_password: str = ""


@dataclass
class CommandTask:
    """命令任务"""

    id: str
    command: str
    description: str = ""
    timeout: int = 30
    require_sudo: bool = False
    dependencies: list[str] = field(default_factory=list)
    status: CommandStatus = CommandStatus.PENDING
    result: dict | None = None
    start_time: datetime | None = None
    end_time: datetime | None = None


class SSHConnection:
    """SSH 连接类"""

    def __init__(self, config: ServerConfig):
        self.config = config
        self.client: paramiko.SSHClient | None = None
        self.sftp: paramiko.SFTPClient | None = None
        self.connected = False
        self.last_activity = time.time()
        self.sudo_session = False

    def connect(self):
        """建立 SSH 连接"""
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            if self.config.use_key and self.config.key_file:
                key_file = os.path.expanduser(self.config.key_file)
                key = None
                for key_class in (
                    paramiko.RSAKey,
                    paramiko.Ed25519Key,
                    paramiko.ECDSAKey,
                ):
                    try:
                        key = key_class.from_private_key_file(key_file)
                        break
                    except (OSError, paramiko.SSHException):
                        continue
                self.client.connect(
                    hostname=self.config.host,
                    port=self.config.port,
                    username=self.config.username,
                    pkey=key,
                    timeout=self.config.timeout,
                    banner_timeout=self.config.timeout,
                    auth_timeout=self.config.timeout,
                    allow_agent=False,
                    look_for_keys=False,
                )
            else:
                self.client.connect(
                    hostname=self.config.host,
                    port=self.config.port,
                    username=self.config.username,
                    password=self.config.password,
                    timeout=self.config.timeout,
                    banner_timeout=self.config.timeout,
                    auth_timeout=self.config.timeout,
                    allow_agent=False,
                    look_for_keys=False,
                )
            self.connected = True
            self.last_activity = time.time()
            return True, "连接成功"
        except Exception as e:
            self.connected = False
            return False, str(e)

    def disconnect(self):
        """断开 SSH 连接"""
        try:
            if self.sftp:
                self.sftp.close()
            if self.client:
                self.client.close()
        except Exception:
            pass
        finally:
            self.sftp = None
            self.client = None
            self.connected = False
            self.sudo_session = False

    def switch_to_sudo(self) -> tuple[bool, str]:
        """切换到 sudo 用户"""
        if not self.connected:
            return False, "未连接"

        try:
            stdin, stdout, stderr = self.client.exec_command(
                f"sudo -s -u {self.config.sudo_user}", timeout=10, get_pty=True
            )
            stdin.write(self.config.sudo_password + "\n")
            stdin.flush()

            exit_code = stdout.channel.recv_exit_status(timeout=10)
            if exit_code == 0:
                self.sudo_session = True
                self.last_activity = time.time()
                return True, f"已切换到用户：{self.config.sudo_user}"
            err = stderr.read().decode("utf-8", errors="replace")
            return False, f"切换失败：{err}"
        except Exception as e:
            return False, f"切换异常：{str(e)}"

    def exec_command(
        self, command: str, timeout: int = None, use_sudo: bool = False
    ) -> tuple[bool, tuple]:
        """执行命令"""
        if not self.client or not self.connected:
            return False, (None, "", "未连接")

        t = timeout if timeout is not None else self.config.timeout

        try:
            if use_sudo and not self.sudo_session:
                command = f"sudo {command}"

            stdin, stdout, stderr = self.client.exec_command(
                command, timeout=t, get_pty=True
            )
            exit_code = stdout.channel.recv_exit_status(timeout=t)
            out = stdout.read().decode("utf-8", errors="replace")
            err = stderr.read().decode("utf-8", errors="replace")

            self.last_activity = time.time()
            return True, (exit_code, out, err)
        except Exception as e:
            self.last_activity = time.time()
            return False, (None, "", str(e))

    def upload_file(self, local_path: str, remote_path: str):
        """上传文件"""
        if not self.client or not self.connected:
            return False, "未连接"
        try:
            if not self.sftp:
                self.sftp = self.client.open_sftp()
            self.sftp.put(local_path, remote_path)
            self.last_activity = time.time()
            return True, "上传成功"
        except Exception as e:
            return False, str(e)

    def download_file(self, remote_path: str, local_path: str):
        """下载文件"""
        if not self.client or not self.connected:
            return False, "未连接"
        try:
            if not self.sftp:
                self.sftp = self.client.open_sftp()
            self.sftp.get(remote_path, local_path)
            self.last_activity = time.time()
            return True, "下载成功"
        except Exception as e:
            return False, str(e)


class SSHPool:
    """SSH 连接池"""

    def __init__(self):
        self.connections: dict[str, SSHConnection] = {}
        self.server_configs: dict[str, ServerConfig] = {}

    def add_server(self, config: ServerConfig) -> str:
        """添加服务器配置"""
        sid = f"{config.host}:{config.port}"
        self.server_configs[sid] = config
        return sid

    def remove_server(self, sid: str):
        """移除服务器"""
        if sid in self.connections:
            self.connections[sid].disconnect()
            del self.connections[sid]
        if sid in self.server_configs:
            del self.server_configs[sid]

    def connect_all(self, progress_callback: Callable = None) -> dict[str, tuple]:
        """连接所有服务器"""
        threads = []
        results = {}

        def _connect(sid):
            if sid not in self.connections:
                conn = SSHConnection(self.server_configs[sid])
                self.connections[sid] = conn
            else:
                conn = self.connections[sid]

            ok, msg = conn.connect()
            results[sid] = (ok, msg)
            if progress_callback:
                progress_callback(sid, ok, msg)

        for sid in self.server_configs:
            t = threading.Thread(target=_connect, args=(sid,), daemon=True)
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        return results

    def disconnect_all(self):
        """断开所有连接"""
        for sid in list(self.connections.keys()):
            self.connections[sid].disconnect()

    def exec_command_all(self, command: str, timeout: int = None) -> dict[str, tuple]:
        """在所有连接上执行命令"""
        threads = []
        results = {}

        def _exec(sid):
            conn = self.connections.get(sid)
            if conn and conn.connected:
                ok, data = conn.exec_command(command, timeout)
                results[sid] = (ok, data)
            else:
                results[sid] = (False, "未连接")

        for sid in self.connections:
            t = threading.Thread(target=_exec, args=(sid,), daemon=True)
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        return results

    def upload_file_all(self, local_path: str, remote_path: str) -> dict[str, tuple]:
        """在所有连接上上传文件"""
        threads = []
        results = {}

        def _upload(sid):
            conn = self.connections.get(sid)
            if conn and conn.connected:
                ok, msg = conn.upload_file(local_path, remote_path)
                results[sid] = (ok, msg)
            else:
                results[sid] = (False, "未连接")

        for sid in self.connections:
            t = threading.Thread(target=_upload, args=(sid,), daemon=True)
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        return results

    def download_file_all(self, remote_path: str, local_path: str) -> dict[str, tuple]:
        """在所有连接上下载文件"""
        threads = []
        results = {}

        def _download(sid):
            conn = self.connections.get(sid)
            if conn and conn.connected:
                ok, msg = conn.download_file(remote_path, local_path)
                results[sid] = (ok, msg)
            else:
                results[sid] = (False, "未连接")

        for sid in self.connections:
            t = threading.Thread(target=_download, args=(sid,), daemon=True)
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        return results

    def get_connected_count(self):
        """获取已连接数量"""
        return sum(1 for c in self.connections.values() if c.connected)


class App:
    """SSH 并行管理工具主程序"""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("SSH 并行管理工具 - 专业版")
        self.root.geometry("1400x900")
        self.root.minsize(1000, 700)

        self.pool = SSHPool()
        self.command_tasks: list[CommandTask] = []
        self._command_confirm_shown = {}
        self._user_config = {}

        self._load_user_config()
        self._build_ui()
        self._load_server_configs()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _get_user_config_path(self):
        """获取用户配置文件路径"""
        app_data = os.environ.get("APPDATA", os.path.expanduser("~"))
        config_dir = os.path.join(app_data, "ssh_parallel_tool")
        os.makedirs(config_dir, exist_ok=True)
        return os.path.join(config_dir, "config.json")

    def _load_user_config(self):
        """加载用户配置"""
        self._user_config = {
            "command_confirm_disabled": {},
            "export_fields": ["name", "host", "port", "username"],
            "default_timeout": 30,
            "execution_mode": CommandExecutionMode.PARALLEL.value,
        }
        config_path = self._get_user_config_path()
        if os.path.exists(config_path):
            try:
                with open(config_path, encoding="utf-8") as f:
                    loaded = json.load(f)
                    self._user_config.update(loaded)
            except Exception:
                pass

    def _save_user_config(self):
        """保存用户配置（原子操作）"""
        config_path = self._get_user_config_path()
        temp_path = config_path + ".tmp"
        try:
            with open(temp_path, "w", encoding="utf-8") as f:
                json.dump(self._user_config, f, ensure_ascii=False, indent=2)
            os.replace(temp_path, config_path)
        except Exception:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    def _show_command_confirm_dialog(
        self,
        command_type: str,
        description: str,
        warning: str,
        command_count: int = 1,
        server_count: int = 1,
    ) -> bool:
        """
        显示命令确认弹窗（单次显示）
        """
        if self._user_config.get("command_confirm_disabled", {}).get(
            command_type, False
        ):
            return True

        if self._command_confirm_shown.get(command_type, False):
            return True

        dialog = tk.Toplevel(self.root)
        dialog.title("命令执行确认")
        dialog.geometry("500x280")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        dialog.bind("<Return>", lambda e: dialog.event_generate("<<Confirm>>"))
        dialog.bind("<Escape>", lambda e: dialog.event_generate("<<Cancel>>"))

        result = {"confirmed": False, "dont_show_again": False}
        dialog_done = threading.Event()

        def on_close():
            result["confirmed"] = False
            dialog_done.set()
            dialog.destroy()

        def on_confirm():
            result["confirmed"] = True
            result["dont_show_again"] = dont_show_var.get()
            dialog_done.set()
            dialog.destroy()

        def on_cancel():
            result["confirmed"] = False
            dialog_done.set()
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", on_close)
        dialog.bind("<<Confirm>>", lambda e: on_confirm())
        dialog.bind("<<Cancel>>", lambda e: on_cancel())

        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        warning_frame = ttk.Frame(main_frame)
        warning_frame.pack(fill=tk.X, pady=(0, 15))

        warning_label = ttk.Label(warning_frame, text="⚠️", font=("Arial", 28))
        warning_label.pack(side=tk.LEFT, padx=(0, 15))

        warning_text = ttk.Label(warning_frame, text=warning, foreground="#FF9800")
        warning_text.pack(side=tk.LEFT, fill=tk.X)

        info_frame = ttk.LabelFrame(main_frame, text="执行信息", padding=12)
        info_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(info_frame, text=f"目标服务器：{server_count} 台").pack(anchor=tk.W)
        ttk.Label(info_frame, text=f"命令数量：{command_count} 条").pack(anchor=tk.W)
        ttk.Label(info_frame, text=description, foreground="#666").pack(
            anchor=tk.W, pady=(8, 0)
        )

        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X)

        dont_show_var = tk.BooleanVar(value=False)
        dont_show_check = ttk.Checkbutton(
            bottom_frame, text="不再显示此提示", variable=dont_show_var
        )
        dont_show_check.pack(side=tk.LEFT)

        button_frame = ttk.Frame(bottom_frame)
        button_frame.pack(side=tk.RIGHT)

        cancel_btn = ttk.Button(button_frame, text="取消", command=on_cancel, width=12)
        cancel_btn.pack(side=tk.LEFT, padx=(0, 12))

        confirm_btn = ttk.Button(
            button_frame, text="确认执行", command=on_confirm, width=12
        )
        confirm_btn.pack(side=tk.LEFT)

        style = ttk.Style()
        style.configure("Confirm.TButton", background="#2196F3", foreground="white")
        style.configure("Cancel.TButton", background="#9E9E9E", foreground="white")
        confirm_btn.configure(style="Confirm.TButton")
        cancel_btn.configure(style="Cancel.TButton")

        dialog.wait_window()
        dialog_done.wait(timeout=5)

        if result["confirmed"]:
            self._command_confirm_shown[command_type] = True

            if result["dont_show_again"]:
                if "command_confirm_disabled" not in self._user_config:
                    self._user_config["command_confirm_disabled"] = {}
                self._user_config["command_confirm_disabled"][command_type] = True
                self._save_user_config()

        return result["confirmed"]

    def _build_ui(self):
        """构建 UI 界面"""
        main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=1)

        right_paned = ttk.PanedWindow(main_paned, orient=tk.VERTICAL)
        main_paned.add(right_paned, weight=2)

        self._build_server_panel(left_frame)
        self._build_command_panel(right_paned)
        self._build_log_panel(right_paned)

        self._build_status_bar()

    def _build_server_panel(self, parent):
        """构建服务器面板"""
        server_frame = ttk.LabelFrame(parent, text="服务器管理", padding=10)
        server_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        tree_frame = ttk.Frame(server_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("name", "host", "port", "username", "status")
        self.server_tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings", height=15
        )
        self.server_tree.heading("name", text="名称")
        self.server_tree.heading("host", text="主机")
        self.server_tree.heading("port", text="端口")
        self.server_tree.heading("username", text="用户名")
        self.server_tree.heading("status", text="状态")

        self.server_tree.column("name", width=100)
        self.server_tree.column("host", width=120)
        self.server_tree.column("port", width=60)
        self.server_tree.column("username", width=80)
        self.server_tree.column("status", width=80)

        scrollbar = ttk.Scrollbar(
            tree_frame, orient=tk.VERTICAL, command=self.server_tree.yview
        )
        self.server_tree.configure(yscrollcommand=scrollbar.set)

        self.server_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        btn_frame = ttk.Frame(server_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(btn_frame, text="添加", command=self._add_server).pack(
            side=tk.LEFT, padx=2
        )
        ttk.Button(btn_frame, text="编辑", command=self._edit_server).pack(
            side=tk.LEFT, padx=2
        )
        ttk.Button(btn_frame, text="删除", command=self._remove_server).pack(
            side=tk.LEFT, padx=2
        )
        ttk.Button(btn_frame, text="全部连接", command=self._connect_all_servers).pack(
            side=tk.LEFT, padx=2
        )
        ttk.Button(
            btn_frame, text="断开所有", command=self._disconnect_all_servers
        ).pack(side=tk.LEFT, padx=2)

        ttk.Button(btn_frame, text="导入", command=self._import_servers).pack(
            side=tk.LEFT, padx=(20, 2)
        )
        ttk.Button(btn_frame, text="导出", command=self._export_servers).pack(
            side=tk.LEFT, padx=2
        )

    def _build_command_panel(self, parent):
        """构建命令面板"""
        cmd_frame = ttk.LabelFrame(parent, text="命令编排与执行", padding=10)
        cmd_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        mode_frame = ttk.Frame(cmd_frame)
        mode_frame.pack(fill=tk.X)

        ttk.Label(mode_frame, text="执行模式:").pack(side=tk.LEFT, padx=(0, 10))

        self.execution_mode_var = tk.StringVar(
            value=self._user_config.get("execution_mode", "parallel")
        )
        ttk.Radiobutton(
            mode_frame,
            text="并行执行",
            variable=self.execution_mode_var,
            value="parallel",
        ).pack(side=tk.LEFT, padx=5)

        ttk.Radiobutton(
            mode_frame,
            text="顺序执行",
            variable=self.execution_mode_var,
            value="sequential",
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            mode_frame, text="保存配置", command=self._save_execution_config
        ).pack(side=tk.RIGHT)

        cmd_text_frame = ttk.Frame(cmd_frame)
        cmd_text_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 10))

        ttk.Label(cmd_text_frame, text="命令列表（每行一条命令，#开头为注释）:").pack(
            anchor=tk.W
        )

        self.cmd_text = scrolledtext.ScrolledText(cmd_text_frame, height=8)
        self.cmd_text.pack(fill=tk.BOTH, expand=True, pady=(5, 0))

        options_frame = ttk.Frame(cmd_frame)
        options_frame.pack(fill=tk.X)

        ttk.Label(options_frame, text="超时时间:").pack(side=tk.LEFT, padx=(0, 5))
        self.cmd_timeout_spin = ttk.Spinbox(options_frame, from_=1, to=300, width=8)
        self.cmd_timeout_spin.set(self._user_config.get("default_timeout", 30))
        self.cmd_timeout_spin.pack(side=tk.LEFT, padx=(0, 20))

        self.sudo_var = tk.BooleanVar(value=False)
        sudo_check = ttk.Checkbutton(
            options_frame, text="使用 sudo 执行", variable=self.sudo_var
        )
        sudo_check.pack(side=tk.LEFT, padx=(0, 20))

        ttk.Button(options_frame, text="执行命令", command=self._exec_command).pack(
            side=tk.RIGHT
        )

        self.cmd_progress = ttk.Progressbar(cmd_frame, mode="determinate")
        self.cmd_progress.pack(fill=tk.X, pady=(10, 5))

        self.cmd_status_label = ttk.Label(cmd_frame, text="就绪")
        self.cmd_status_label.pack(anchor=tk.W)

    def _build_log_panel(self, parent):
        """构建日志面板"""
        log_frame = ttk.LabelFrame(parent, text="执行日志", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=12)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        self.log_text.tag_configure("info", foreground="#2196F3")
        self.log_text.tag_configure("success", foreground="#4CAF50")
        self.log_text.tag_configure("error", foreground="#F44336")
        self.log_text.tag_configure("warning", foreground="#FF9800")
        self.log_text.tag_configure("header", foreground="#673AB7")

        log_btn_frame = ttk.Frame(log_frame)
        log_btn_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Button(log_btn_frame, text="清空日志", command=self._clear_log).pack(
            side=tk.LEFT
        )
        ttk.Button(log_btn_frame, text="导出日志", command=self._export_log).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Button(log_btn_frame, text="保存日志", command=self._save_log).pack(
            side=tk.LEFT
        )

    def _build_status_bar(self):
        """构建状态栏"""
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)

        self.status_label = ttk.Label(
            status_frame, text="就绪 | 已连接：0/0", relief=tk.SUNKEN, padding=5
        )
        self.status_label.pack(fill=tk.X, padx=5, pady=5)

    def _log_output(self, message: str, level: str = "info"):
        """输出日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted = f"[{timestamp}] {message}\n"

        try:
            self.log_text.insert(tk.END, formatted, level)
            self.log_text.see(tk.END)
        except Exception:
            pass

    def _clear_log(self):
        """清空日志"""
        self.log_text.delete(1.0, tk.END)
        self._log_output("日志已清空", "info")

    def _save_log(self):
        """保存日志"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            title="保存日志",
        )
        if file_path:
            try:
                content = self.log_text.get(1.0, tk.END)
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(content)
                messagebox.showinfo("成功", "日志已保存")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败：{str(e)}")

    def _export_log(self):
        """导出日志"""
        self._save_log()

    def _update_server_tree(self):
        """更新服务器树"""
        for item in self.server_tree.get_children():
            self.server_tree.delete(item)

        for sid, config in self.pool.server_configs.items():
            conn = self.pool.connections.get(sid)
            status = "✓ 已连接" if conn and conn.connected else "○ 未连接"
            display_name = config.name if config.name else config.host
            self.server_tree.insert(
                "",
                tk.END,
                values=(
                    display_name,
                    config.host,
                    config.port,
                    config.username,
                    status,
                ),
            )

        connected = self.pool.get_connected_count()
        total = len(self.pool.server_configs)
        self.status_label.config(text=f"就绪 | 已连接：{connected}/{total}")
        self.root.update()

    def _add_server(self):
        """添加服务器"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加服务器")
        dialog.geometry("450x350")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        result = {}

        def on_save():
            result["saved"] = True
            dialog.destroy()

        def on_cancel():
            result["saved"] = False
            dialog.destroy()

        fields_frame = ttk.Frame(dialog, padding=20)
        fields_frame.pack(fill=tk.BOTH, expand=True)

        row = 0
        ttk.Label(fields_frame, text="名称:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        name_entry = ttk.Entry(fields_frame, width=40)
        name_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="主机:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        host_entry = ttk.Entry(fields_frame, width=40)
        host_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="端口:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        port_spin = ttk.Spinbox(fields_frame, from_=1, to=65535, width=10)
        port_spin.set(22)
        port_spin.grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1

        ttk.Label(fields_frame, text="用户名:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        user_entry = ttk.Entry(fields_frame, width=40)
        user_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="密码:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        pass_entry = ttk.Entry(fields_frame, width=40, show="*")
        pass_entry.grid(row=row, column=1, pady=5)
        row += 1

        use_sudo_var = tk.BooleanVar(value=False)
        ttk.Label(fields_frame, text="使用 Sudo:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        sudo_check = ttk.Checkbutton(fields_frame, variable=use_sudo_var)
        sudo_check.grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1

        ttk.Label(fields_frame, text="Sudo 用户:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        sudo_user_entry = ttk.Entry(fields_frame, width=40)
        sudo_user_entry.insert(0, "root")
        sudo_user_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="Sudo 密码:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        sudo_pass_entry = ttk.Entry(fields_frame, width=40, show="*")
        sudo_pass_entry.grid(row=row, column=1, pady=5)
        row += 1

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(btn_frame, text="保存", command=on_save).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.RIGHT)

        dialog.wait_window()

        if result.get("saved"):
            config = ServerConfig(
                name=name_entry.get().strip(),
                host=host_entry.get().strip(),
                port=int(port_spin.get()),
                username=user_entry.get().strip(),
                password=pass_entry.get(),
                use_sudo=use_sudo_var.get(),
                sudo_user=sudo_user_entry.get().strip() or "root",
                sudo_password=sudo_pass_entry.get(),
            )

            if config.host:
                self.pool.add_server(config)
                self._update_server_tree()
                self._log_output(f"添加服务器：{config.host}", "success")

    def _edit_server(self):
        """编辑服务器"""
        selected = self.server_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请选择要编辑的服务器")
            return

        item = self.server_tree.item(selected[0])
        values = item["values"]
        host, port = values[1], values[2]
        sid = f"{host}:{port}"

        if sid not in self.pool.server_configs:
            return

        config = self.pool.server_configs[sid]

        dialog = tk.Toplevel(self.root)
        dialog.title("编辑服务器")
        dialog.geometry("450x350")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        result = {}

        def on_save():
            result["saved"] = True
            dialog.destroy()

        def on_cancel():
            result["saved"] = False
            dialog.destroy()

        fields_frame = ttk.Frame(dialog, padding=20)
        fields_frame.pack(fill=tk.BOTH, expand=True)

        row = 0
        ttk.Label(fields_frame, text="名称:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        name_entry = ttk.Entry(fields_frame, width=40)
        name_entry.insert(0, config.name)
        name_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="主机:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        host_entry = ttk.Entry(fields_frame, width=40)
        host_entry.insert(0, config.host)
        host_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="端口:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        port_spin = ttk.Spinbox(fields_frame, from_=1, to=65535, width=10)
        port_spin.set(config.port)
        port_spin.grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1

        ttk.Label(fields_frame, text="用户名:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        user_entry = ttk.Entry(fields_frame, width=40)
        user_entry.insert(0, config.username)
        user_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="密码:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        pass_entry = ttk.Entry(fields_frame, width=40, show="*")
        pass_entry.insert(0, config.password)
        pass_entry.grid(row=row, column=1, pady=5)
        row += 1

        use_sudo_var = tk.BooleanVar(value=config.use_sudo)
        ttk.Label(fields_frame, text="使用 Sudo:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        sudo_check = ttk.Checkbutton(fields_frame, variable=use_sudo_var)
        sudo_check.grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1

        ttk.Label(fields_frame, text="Sudo 用户:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        sudo_user_entry = ttk.Entry(fields_frame, width=40)
        sudo_user_entry.insert(0, config.sudo_user)
        sudo_user_entry.grid(row=row, column=1, pady=5)
        row += 1

        ttk.Label(fields_frame, text="Sudo 密码:").grid(
            row=row, column=0, sticky=tk.W, pady=5
        )
        sudo_pass_entry = ttk.Entry(fields_frame, width=40, show="*")
        sudo_pass_entry.grid(row=row, column=1, pady=5)
        row += 1

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(btn_frame, text="保存", command=on_save).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.RIGHT)

        dialog.wait_window()

        if result.get("saved"):
            new_config = ServerConfig(
                name=name_entry.get().strip(),
                host=host_entry.get().strip(),
                port=int(port_spin.get()),
                username=user_entry.get().strip(),
                password=pass_entry.get(),
                use_sudo=use_sudo_var.get(),
                sudo_user=sudo_user_entry.get().strip() or "root",
                sudo_password=sudo_pass_entry.get(),
            )
            self.pool.server_configs[sid] = new_config
            self._update_server_tree()
            self._log_output(f"更新服务器：{new_config.host}", "success")

    def _remove_server(self):
        """删除服务器"""
        selected = self.server_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请选择要删除的服务器")
            return

        if messagebox.askyesno("确认删除", "确定要删除选中的服务器吗？"):
            for item in selected:
                values = self.server_tree.item(item)["values"]
                host, port = values[1], values[2]
                sid = f"{host}:{port}"
                self.pool.remove_server(sid)
            self._update_server_tree()
            self._log_output("服务器已删除", "info")

    def _connect_all_servers(self):
        """连接所有服务器"""
        if not self.pool.server_configs:
            messagebox.showinfo("提示", "没有服务器需要连接")
            return

        self._log_output("正在连接所有服务器...", "info")

        def _do_connect():
            def progress(sid, ok, msg):
                self.root.after(0, lambda: self._on_connect_progress(sid, ok, msg))

            self.pool.connect_all(progress_callback=progress)
            self.root.after(0, self._update_server_tree)
            self.root.after(
                0, lambda: self._log_output("所有服务器连接完成", "success")
            )

        threading.Thread(target=_do_connect, daemon=True).start()

    def _on_connect_progress(self, sid, ok, msg):
        """连接进度回调"""
        tag = "success" if ok else "error"
        self._log_output(f"[{sid}] {'连接成功' if ok else '连接失败：' + msg}", tag)
        self._update_server_tree()

    def _disconnect_all_servers(self):
        """断开所有服务器"""
        self.pool.disconnect_all()
        self._update_server_tree()
        self._log_output("已断开所有服务器连接", "info")

    def _exec_command(self):
        """执行命令"""
        raw_commands = self.cmd_text.get(1.0, tk.END).strip()
        if not raw_commands:
            messagebox.showwarning("提示", "请输入要执行的命令")
            return

        commands = []
        for line in raw_commands.split("\n"):
            line = line.strip()
            if line and not line.startswith("#"):
                commands.append(line)

        if not commands:
            messagebox.showwarning("提示", "没有有效的命令")
            return

        connected = [sid for sid, c in self.pool.connections.items() if c.connected]
        if not connected:
            messagebox.showwarning("提示", "没有已连接的服务器")
            return

        use_sudo = self.sudo_var.get()
        timeout = int(self.cmd_timeout_spin.get())

        confirmed = self._show_command_confirm_dialog(
            "batch_command_exec",
            f"即将在 {len(connected)} 台服务器上执行 {len(commands)} 条命令",
            "警告：此操作将在所有已连接服务器上执行命令，请确保命令安全！",
            len(commands),
            len(connected),
        )
        if not confirmed:
            return

        execution_mode = self.execution_mode_var.get()
        self._log_output(
            f"开始执行命令 (模式：{execution_mode}, 超时：{timeout}s)", "header"
        )

        total = len(connected)
        self.cmd_progress["maximum"] = total
        self.cmd_progress["value"] = 0
        self.cmd_status_label.config(text="执行中...")

        def _do_exec():
            if execution_mode == "parallel":
                results = self._exec_commands_parallel(
                    commands, connected, timeout, use_sudo
                )
            else:
                results = self._exec_commands_sequential(
                    commands, connected, timeout, use_sudo
                )

            completed = 0
            for sid, (ok, data) in results.items():
                completed += 1
                if ok:
                    exit_code, stdout, stderr = data
                    self.root.after(
                        0,
                        lambda s=sid, ec=exit_code, so=stdout, se=stderr: (
                            self._on_cmd_result(s, ec, so, se)
                        ),
                    )
                else:
                    self.root.after(
                        0,
                        lambda s=sid, e=data: self._log_output(
                            f"[{s}] 执行失败：{e}", "error"
                        ),
                    )
                self.root.after(
                    0, lambda v=completed: self.cmd_progress.configure(value=v)
                )

            self.root.after(0, lambda: self.cmd_status_label.config(text="执行完成"))
            self.root.after(
                0,
                lambda: self._log_output(
                    f"命令执行完毕 ({completed}/{total})", "success"
                ),
            )

        threading.Thread(target=_do_exec, daemon=True).start()

    def _exec_commands_parallel(self, commands, servers, timeout, use_sudo):
        """并行执行命令"""
        results = {}
        threads = []

        def _exec(sid, cmd_list):
            conn = self.pool.connections.get(sid)
            if not conn or not conn.connected:
                results[sid] = (False, "未连接")
                return

            for cmd in cmd_list:
                ok, data = conn.exec_command(cmd, timeout, use_sudo=use_sudo)
                if not ok:
                    results[sid] = (False, data)
                    return

            results[sid] = (True, data)

        for sid in servers:
            t = threading.Thread(target=_exec, args=(sid, commands), daemon=True)
            threads.append(t)
            t.start()

        for t in threads:
            t.join()

        return results

    def _exec_commands_sequential(self, commands, servers, timeout, use_sudo):
        """顺序执行命令"""
        results = {}

        for sid in servers:
            conn = self.pool.connections.get(sid)
            if not conn or not conn.connected:
                results[sid] = (False, "未连接")
                continue

            success = True
            last_data = None
            for cmd in commands:
                if not success:
                    break
                ok, data = conn.exec_command(cmd, timeout, use_sudo=use_sudo)
                if not ok:
                    success = False
                last_data = data

            results[sid] = (success, last_data)

        return results

    def _on_cmd_result(self, sid, exit_code, stdout, stderr):
        """命令执行结果回调"""
        header = f"═══════ [{sid}] 退出码：{exit_code} ═══════"
        self._log_output(header, "header")

        if stdout.strip():
            self._log_output(f"输出:\n{stdout}", "success")

        if stderr.strip():
            self._log_output(f"错误:\n{stderr}", "error")

    def _import_servers(self):
        """导入服务器配置"""
        file_path = filedialog.askopenfilename(
            title="导入服务器配置",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("JSON 文件", "*.json"),
                ("所有文件", "*.*"),
            ],
        )
        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx"):
                self._import_from_excel(file_path)
            elif file_path.endswith(".json"):
                self._import_from_json(file_path)
            else:
                messagebox.showerror("错误", "不支持的文件格式")
                return

            self._update_server_tree()
            messagebox.showinfo(
                "成功", f"成功导入 {len(self.pool.server_configs)} 个服务器配置"
            )
            self._log_output(f"从 {file_path} 导入服务器配置", "success")
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{str(e)}")
            self._log_output(f"导入失败：{str(e)}", "error")

    def _import_from_excel(self, file_path):
        """从 Excel 导入"""
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers[cell_value.lower()] = col

        required = {"host"}
        if not required.issubset(set(headers.keys())):
            raise ValueError("Excel 文件缺少必需的列：host")

        for row in range(2, ws.max_row + 1):
            host = ws.cell(row=row, column=headers.get("host", 0)).value
            if not host:
                continue

            name = (
                ws.cell(row=row, column=headers.get("name", 0)).value
                if "name" in headers
                else ""
            )
            port = (
                ws.cell(row=row, column=headers.get("port", 0)).value
                if "port" in headers
                else 22
            )
            username = (
                ws.cell(row=row, column=headers.get("username", 0)).value
                if "username" in headers
                else ""
            )
            password = (
                ws.cell(row=row, column=headers.get("password", 0)).value
                if "password" in headers
                else ""
            )

            config = ServerConfig(
                name=str(name) if name else str(host),
                host=str(host),
                port=int(port) if port else 22,
                username=str(username) if username else "",
                password=str(password) if password else "",
            )
            self.pool.add_server(config)

        wb.close()

    def _import_from_json(self, file_path):
        """从 JSON 导入"""
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            for item in data:
                config = ServerConfig(
                    name=item.get("name", item.get("host", "")),
                    host=item.get("host", ""),
                    port=item.get("port", 22),
                    username=item.get("username", ""),
                    password=item.get("password", ""),
                )
                if config.host:
                    self.pool.add_server(config)

    def _export_servers(self):
        """导出服务器配置"""
        if not self.pool.server_configs:
            messagebox.showwarning("提示", "没有服务器配置可导出")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("JSON 文件", "*.json"),
                ("所有文件", "*.*"),
            ],
            title="导出服务器配置",
        )
        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx"):
                self._export_to_excel(file_path)
            elif file_path.endswith(".json"):
                self._export_to_json(file_path)
            else:
                messagebox.showerror("错误", "不支持的文件格式")
                return

            messagebox.showinfo(
                "成功", f"成功导出 {len(self.pool.server_configs)} 个服务器配置"
            )
            self._log_output(f"导出服务器配置到 {file_path}", "success")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")
            self._log_output(f"导出失败：{str(e)}", "error")

    def _export_to_excel(self, file_path):
        """导出到 Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Servers"

        headers = ["名称", "主机", "端口", "用户名", "密码"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")

        for row, (sid, config) in enumerate(self.pool.server_configs.items(), 2):
            ws.cell(row=row, column=1, value=config.name)
            ws.cell(row=row, column=2, value=config.host)
            ws.cell(row=row, column=3, value=config.port)
            ws.cell(row=row, column=4, value=config.username)
            ws.cell(row=row, column=5, value=config.password)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        wb.save(file_path)
        wb.close()

    def _export_to_json(self, file_path):
        """导出到 JSON"""
        data = []
        for sid, config in self.pool.server_configs.items():
            data.append(
                {
                    "name": config.name,
                    "host": config.host,
                    "port": config.port,
                    "username": config.username,
                    "password": config.password,
                }
            )

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _save_execution_config(self):
        """保存执行配置"""
        self._user_config["execution_mode"] = self.execution_mode_var.get()
        self._user_config["default_timeout"] = int(self.cmd_timeout_spin.get())
        self._save_user_config()
        self._log_output("执行配置已保存", "success")

    def _load_server_configs(self):
        """加载服务器配置（示例）"""
        pass

    def _on_close(self):
        """关闭应用"""
        if messagebox.askyesno("确认退出", "确定要退出吗？"):
            self.pool.disconnect_all()
            self.root.destroy()


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
